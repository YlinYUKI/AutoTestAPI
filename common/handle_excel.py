"""
    1、数据读取封装一个可以读取任意excel文件的方法，可以指定读取的表单
    2、数据写入
            文件名：
            表单：
            行：
            列：
            写入的值：

"""
import openpyxl

class HandleExcel:
    #初始化
    def __init__(self,filename,sheetname):
        """
        :param filename: excel文件名（同一个目录）  excel文件的绝对路径（不在一个目录下）
        :param sheetname: （表单名字）
        """
        self.filename = filename
        self.sheetname = sheetname


    """读取excel数据"""
    def read_excel(self):
        # 读取工作簿
        workbook = openpyxl.load_workbook(self.filename)
        # 指定的表单
        sh = workbook[self.sheetname]
        # rows 接收表单中的所有数据
        res = list(sh.rows)
        # 拿到第一行的数据
        title = [i.value for i in res[0]]
        # 定义一个空列表来存储字典
        cases = []
        # 遍历除第一行以外的数据
        for item in res[1:]:
            # 读取每一行的数据
            data = [i.value for i in item]
            # 组装成一个字典
            dic = dict(zip(title, data))
            cases.append(dic)
        #返回读取出来的数据
        return cases


    """数据写入的方法"""
    def write_excel(self, row, column, value):
        """
        数据写入方法
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        """
        #加载工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # excel中写入数据
        sh.cell(row=row, column=column, value=value)
        # 写入后要保存
        workbook.save(self.filename)